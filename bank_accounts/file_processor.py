"""
Enhanced Bank Statement File Processor
Handles Excel and CSV files with 100% original data preservation
"""

import hashlib
import json
import logging
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Tuple, Optional, Any
import uuid

import openpyxl
import csv
from django.core.files.uploadedfile import UploadedFile as DjangoUploadedFile
from django.db import transaction
from django.utils import timezone

from .models import (
    BankStatementDocument,
    BankStatementProcessing,
    BankTransactionRecord,
)
from .translation_service import MongolianTranslationService
from coa.models import Account

logger = logging.getLogger(__name__)


class BankStatementProcessor:
    """
    Processes bank statement files with immutable data preservation
    and translation services
    """

    SUPPORTED_FORMATS = {"xlsx": ["xlsx", "xls"], "csv": ["csv", "txt"]}

    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB limit

    def __init__(self, company, user=None):
        self.company = company
        self.user = user
        self.translation_service = MongolianTranslationService()

    def process_file(
        self, uploaded_file: DjangoUploadedFile, account_id: Optional[int] = None
    ) -> Dict:
        """
        Main processing entry point
        Returns: processing results summary
        """
        try:
            # Validate file
            validation_result = self._validate_file(uploaded_file)
            if not validation_result["valid"]:
                return validation_result

            # Determine file format
            file_format = self._get_file_format(uploaded_file.name)

            # Read and preserve original data
            raw_data = self._read_file_contents(uploaded_file, file_format)
            if not raw_data["success"]:
                return raw_data

            # Create immutable document record
            document = self._create_document_record(
                uploaded_file, file_format, raw_data["data"]
            )

            # Process and translate data
            processing_result = self._process_document_data(document, account_id)

            return {
                "success": True,
                "document_id": str(document.document_id),
                "processing_id": str(processing_result.processing_id),
                "summary": self._generate_processing_summary(processing_result),
            }

        except Exception as e:
            logger.error(f"File processing error: {str(e)}")
            return {
                "success": False,
                "error": f"Processing failed: {str(e)}",
                "error_type": "processing_error",
            }

    def _validate_file(self, uploaded_file: DjangoUploadedFile) -> Dict:
        """Validate uploaded file"""

        # Check file size
        if uploaded_file.size > self.MAX_FILE_SIZE:
            return {
                "valid": False,
                "error": f"File too large. Maximum size is {self.MAX_FILE_SIZE // (1024*1024)}MB",
                "error_type": "file_too_large",
            }

        # Check file format
        file_format = self._get_file_format(uploaded_file.name)
        if not file_format:
            return {
                "valid": False,
                "error": "Unsupported file format. Please upload Excel (.xlsx) or CSV (.csv) files",
                "error_type": "unsupported_format",
            }

        # Check if file has content
        if uploaded_file.size == 0:
            return {
                "valid": False,
                "error": "File is empty",
                "error_type": "empty_file",
            }

        return {"valid": True}

    def _get_file_format(self, filename: str) -> Optional[str]:
        """Determine file format from extension"""
        if not filename:
            return None

        extension = filename.lower().split(".")[-1]

        for format_type, extensions in self.SUPPORTED_FORMATS.items():
            if extension in extensions:
                return format_type

        return None

    def _read_file_contents(
        self, uploaded_file: DjangoUploadedFile, file_format: str
    ) -> Dict:
        """Read and preserve 100% original file contents"""

        try:
            if file_format == "xlsx":
                return self._read_excel_file(uploaded_file)
            elif file_format == "csv":
                return self._read_csv_file(uploaded_file)
            else:
                return {"success": False, "error": f"Unsupported format: {file_format}"}
        except Exception as e:
            logger.error(f"File reading error: {str(e)}")
            return {"success": False, "error": f"Could not read file: {str(e)}"}

    def _read_excel_file(self, uploaded_file: DjangoUploadedFile) -> Dict:
        """Read Excel file preserving all data"""

        try:
            # Load workbook
            workbook = openpyxl.load_workbook(uploaded_file, data_only=False)
            sheet = workbook.active

            raw_data = {
                "header_rows": [],
                "transaction_rows": [],
                "sheet_info": {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                },
            }

            # Read all rows preserving exact content
            all_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                row_data = []
                for cell_value in row:
                    # Convert to string preserving original formatting
                    if cell_value is None:
                        row_data.append("")
                    elif isinstance(cell_value, datetime):
                        row_data.append(cell_value.strftime("%Y-%m-%d %H:%M:%S"))
                    else:
                        row_data.append(str(cell_value))

                all_rows.append({"row_number": row_idx, "cells": row_data})

            # Detect header vs transaction rows
            header_end_row = self._detect_header_section(all_rows)

            raw_data["header_rows"] = all_rows[:header_end_row]
            raw_data["transaction_rows"] = all_rows[header_end_row:]

            return {"success": True, "data": raw_data}

        except Exception as e:
            logger.error(f"Excel reading error: {str(e)}")
            return {"success": False, "error": f"Excel file error: {str(e)}"}

    def _read_csv_file(self, uploaded_file: DjangoUploadedFile) -> Dict:
        """Read CSV file preserving all data"""

        try:
            # Try different encodings
            encodings = ["utf-8", "utf-8-sig", "cp1252", "iso-8859-1"]
            content = None
            used_encoding = None

            for encoding in encodings:
                try:
                    uploaded_file.seek(0)
                    content = uploaded_file.read().decode(encoding)
                    used_encoding = encoding
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:
                return {
                    "success": False,
                    "error": "Could not decode file with any supported encoding",
                }

            # Parse CSV
            csv_reader = csv.reader(content.splitlines())
            all_rows = []

            for row_idx, row in enumerate(csv_reader, 1):
                all_rows.append({"row_number": row_idx, "cells": row})

            # For CSV, assume first row is header
            raw_data = {
                "header_rows": all_rows[:1] if all_rows else [],
                "transaction_rows": all_rows[1:] if len(all_rows) > 1 else [],
                "encoding_used": used_encoding,
            }

            return {"success": True, "data": raw_data}

        except Exception as e:
            logger.error(f"CSV reading error: {str(e)}")
            return {"success": False, "error": f"CSV file error: {str(e)}"}

    def _detect_header_section(self, all_rows: List[Dict]) -> int:
        """
        Detect where header section ends and transaction data begins
        Returns: row index where transactions start (0-based)
        """

        # Look for patterns that indicate start of transaction data
        for idx, row_data in enumerate(all_rows):
            cells = row_data.get("cells", [])

            # Skip empty rows
            if not any(str(cell).strip() for cell in cells):
                continue

            # Look for column headers (Date, Amount, etc.)
            headers_found = 0
            for cell in cells:
                cell_str = str(cell).strip().lower()
                if any(
                    header in cell_str
                    for header in [
                        "огноо",
                        "дүн",
                        "тайлбар",
                        "date",
                        "amount",
                        "description",
                    ]
                ):
                    headers_found += 1

            # If we found multiple column headers, next row likely starts transactions
            if headers_found >= 2:
                return idx + 1

        # Default: assume first 10 rows are header if no clear pattern
        return min(10, len(all_rows))

    @transaction.atomic
    def _create_document_record(
        self, uploaded_file: DjangoUploadedFile, file_format: str, raw_data: Dict
    ) -> BankStatementDocument:
        """Create immutable document record"""

        # Calculate file hash
        uploaded_file.seek(0)
        file_content = uploaded_file.read()
        file_hash = hashlib.sha256(file_content).hexdigest()

        # Create document record
        document = BankStatementDocument.objects.create(
            original_filename=uploaded_file.name,
            file_format=file_format,
            file_size_bytes=uploaded_file.size,
            uploaded_by=self.user,
            company=self.company,
            file_hash=file_hash,
            raw_header_data=raw_data.get("header_rows", []),
            raw_transaction_data=raw_data.get("transaction_rows", []),
            original_encoding=raw_data.get("encoding_used", "utf-8"),
        )

        return document

    @transaction.atomic
    def _process_document_data(
        self, document: BankStatementDocument, account_id: Optional[int] = None
    ) -> BankStatementProcessing:
        """Process document data with translations"""

        # Detect overall language
        sample_text = self._extract_sample_text(document)
        detected_language, language_confidence = (
            self.translation_service.detect_language(sample_text)
        )

        # Create processing record
        processing = BankStatementProcessing.objects.create(
            source_document=document,
            processed_by=self.user,
            detected_language=detected_language,
            language_confidence=language_confidence,
        )

        # Process header information
        header_info = self._process_header_data(document.raw_header_data, processing)

        # Update processing record with header info
        processing.system_account_holder = header_info.get("account_holder", "")
        processing.system_account_number = header_info.get("account_number", "")
        processing.system_account_type = header_info.get("account_type", "")
        processing.system_date_range_start = header_info.get("date_start")
        processing.system_date_range_end = header_info.get("date_end")
        processing.translation_details = header_info.get("translation_details", {})

        # Try to link with existing account
        if account_id:
            try:
                account = Account.objects.get(id=account_id, company=self.company)
                processing.suggested_coa_account = account
                processing.auto_link_confidence = Decimal(
                    "0.8"
                )  # Manual selection gets high confidence
            except Account.DoesNotExist:
                pass
        else:
            # Try auto-linking based on account number
            account_number = header_info.get("account_number", "")
            if account_number:
                matching_accounts = Account.objects.filter(
                    company=self.company, code__icontains=account_number
                ).filter(is_bank_account=True)

                if matching_accounts.count() == 1:
                    processing.suggested_coa_account = matching_accounts.first()
                    processing.auto_link_confidence = Decimal("0.7")

        # Process transaction data
        self._process_transaction_data(document.raw_transaction_data, processing)

        # Update final statistics
        transaction_count = processing.transaction_records.count()
        processing.total_transactions = transaction_count
        processing.successfully_processed = processing.transaction_records.filter(
            overall_confidence__gte=0.5
        ).count()
        processing.failed_translations = (
            transaction_count - processing.successfully_processed
        )
        processing.processing_complete = True
        processing.save()

        return processing

    def _extract_sample_text(self, document: BankStatementDocument) -> str:
        """Extract sample text for language detection"""

        sample_parts = []

        # Sample from headers
        for header_row in document.raw_header_data[:5]:
            cells = header_row.get("cells", [])
            for cell in cells[:3]:  # First 3 cells of each row
                if isinstance(cell, str) and cell.strip():
                    sample_parts.append(cell.strip())

        # Sample from transactions
        for trans_row in document.raw_transaction_data[:10]:
            cells = trans_row.get("cells", [])
            for cell in cells:
                if isinstance(cell, str) and cell.strip() and len(cell) > 3:
                    sample_parts.append(cell.strip())
                    break  # One sample per transaction row

        return " ".join(sample_parts[:50])  # Limit sample size

    def _process_header_data(
        self, header_rows: List[Dict], processing: BankStatementProcessing
    ) -> Dict:
        """Process header information with translations"""

        header_info = {
            "account_holder": "",
            "account_number": "",
            "account_type": "",
            "date_start": None,
            "date_end": None,
            "translation_details": {},
        }

        # Process each header row
        for row_data in header_rows:
            cells = row_data.get("cells", [])

            for i, cell in enumerate(cells):
                if not isinstance(cell, str) or not cell.strip():
                    continue

                cell_content = cell.strip()

                # Check for known header patterns
                if "Хэрэглэгч" in cell_content or "Account Holder" in cell_content:
                    # Next cell likely contains account holder name
                    if i + 1 < len(cells) and cells[i + 1]:
                        header_info["account_holder"] = str(cells[i + 1]).strip()

                elif (
                    "Дансны дугаар" in cell_content or "Account Number" in cell_content
                ):
                    # Extract account number
                    if i + 1 < len(cells) and cells[i + 1]:
                        account_info = str(cells[i + 1]).strip()
                        # Extract just the number part
                        import re

                        numbers = re.findall(r"\d{8,}", account_info)
                        if numbers:
                            header_info["account_number"] = numbers[0]

                        # Extract account type
                        type_match = re.search(r"\((.*?)\)", account_info)
                        if type_match:
                            account_type_raw = type_match.group(1)
                            type_translation = (
                                self.translation_service.translate_account_type(
                                    account_type_raw
                                )
                            )
                            header_info["account_type"] = type_translation["english"]

                elif "Интервал" in cell_content or "Date Range" in cell_content:
                    # Extract date range
                    if i + 1 < len(cells) and cells[i + 1]:
                        date_range = str(cells[i + 1]).strip()
                        dates = self._parse_date_range(date_range)
                        header_info["date_start"] = dates[0]
                        header_info["date_end"] = dates[1]

        return header_info

    def _parse_date_range(
        self, date_range_text: str
    ) -> Tuple[Optional[datetime], Optional[datetime]]:
        """Parse date range like '2025/04/01 - 2025/06/30'"""

        import re

        # Look for date range patterns
        date_pattern = r"(\d{4}/\d{1,2}/\d{1,2})\s*[-–]\s*(\d{4}/\d{1,2}/\d{1,2})"
        match = re.search(date_pattern, date_range_text)

        if match:
            start_date_str = match.group(1)
            end_date_str = match.group(2)

            start_date, _ = self.translation_service.parse_date(start_date_str)
            end_date, _ = self.translation_service.parse_date(end_date_str)

            return start_date, end_date

        return None, None

    def _process_transaction_data(
        self, transaction_rows: List[Dict], processing: BankStatementProcessing
    ):
        """Process transaction data with translations"""

        if not transaction_rows:
            return

        # Detect column headers from first row
        column_mapping = self._detect_column_mapping(
            transaction_rows[0] if transaction_rows else {}
        )

        # Process each transaction row
        for row_data in transaction_rows[1:]:  # Skip header row
            cells = row_data.get("cells", [])
            row_number = row_data.get("row_number", 0)

            # Skip empty rows
            if not any(str(cell).strip() for cell in cells):
                continue

            # Extract data based on column mapping
            transaction_data = self._extract_transaction_data(cells, column_mapping)

            # Create transaction record
            self._create_transaction_record(transaction_data, row_number, processing)

    def _detect_column_mapping(self, header_row: Dict) -> Dict[str, int]:
        """Detect which columns contain which data"""

        mapping = {}
        cells = header_row.get("cells", [])

        for i, cell in enumerate(cells):
            if not isinstance(cell, str):
                continue

            cell_lower = cell.strip().lower()

            # Map known headers to column indices
            if "огноо" in cell_lower or "date" in cell_lower:
                mapping["date"] = i
            elif "дүн" in cell_lower or "amount" in cell_lower:
                mapping["amount"] = i
            elif "тайлбар" in cell_lower or "description" in cell_lower:
                mapping["description"] = i
            elif "лавлагаа" in cell_lower or "reference" in cell_lower:
                mapping["reference"] = i
            elif "хүлээн авагч" in cell_lower or "payee" in cell_lower:
                mapping["payee"] = i
            elif (
                "харьцсан данс" in cell_lower
                or "related_account" in cell_lower
                or "related account" in cell_lower
            ):
                mapping["related_account"] = i

        return mapping

    def _extract_transaction_data(
        self, cells: List, column_mapping: Dict[str, int]
    ) -> Dict:
        """Extract transaction data from row cells"""

        data = {
            "original_date_text": "",
            "original_amount_text": "",
            "original_description": "",
            "original_payee": "",
            "original_reference": "",
            "original_related_account": "",
        }

        for field, column_index in column_mapping.items():
            if column_index < len(cells) and cells[column_index]:
                value = str(cells[column_index]).strip()
                if field in ["date", "amount"]:
                    data[f"original_{field}_text"] = value
                else:
                    data[f"original_{field}"] = value

        return data

    def _create_transaction_record(
        self,
        transaction_data: Dict,
        row_number: int,
        processing: BankStatementProcessing,
    ):
        """Create transaction record with translations"""

        # Parse and translate data
        system_date, date_confidence = self.translation_service.parse_date(
            transaction_data.get("original_date_text", "")
        )

        system_amount, amount_confidence = self.translation_service.parse_amount(
            transaction_data.get("original_amount_text", "")
        )

        description_translation = self.translation_service.translate_description(
            transaction_data.get("original_description", "")
        )

        # Calculate overall confidence
        confidences = [
            date_confidence,
            amount_confidence,
            description_translation["confidence"],
        ]
        overall_confidence = (
            sum(c for c in confidences if c > 0)
            / len([c for c in confidences if c > 0])
            if any(confidences)
            else 0.0
        )

        # Create record
        BankTransactionRecord.objects.create(
            processing=processing,
            row_number=row_number,
            original_date_text=transaction_data.get("original_date_text", ""),
            original_amount_text=transaction_data.get("original_amount_text", ""),
            original_description=transaction_data.get("original_description", ""),
            original_payee=transaction_data.get("original_payee", ""),
            original_reference=transaction_data.get("original_reference", ""),
            system_date=system_date,
            system_amount=system_amount,
            system_description_english=description_translation["english"],
            date_parse_confidence=Decimal(str(date_confidence)),
            amount_parse_confidence=Decimal(str(amount_confidence)),
            description_translation_confidence=Decimal(
                str(description_translation["confidence"])
            ),
            overall_confidence=Decimal(str(overall_confidence)),
            translation_notes={
                "terms_found": description_translation.get("terms_found", []),
                "processing_timestamp": timezone.now().isoformat(),
            },
        )

    def _generate_processing_summary(self, processing: BankStatementProcessing) -> Dict:
        """Generate processing summary for UI"""

        records = processing.transaction_records.all()

        return {
            "total_transactions": records.count(),
            "high_confidence": records.filter(overall_confidence__gte=0.8).count(),
            "medium_confidence": records.filter(
                overall_confidence__gte=0.5, overall_confidence__lt=0.8
            ).count(),
            "low_confidence": records.filter(overall_confidence__lt=0.5).count(),
            "detected_language": processing.get_detected_language_display(),
            "language_confidence": float(processing.language_confidence),
            "account_linked": processing.suggested_coa_account is not None,
            "needs_verification": records.filter(overall_confidence__lt=0.8).count(),
            "header_info": {
                "account_holder": processing.system_account_holder,
                "account_number": processing.system_account_number,
                "account_type": processing.system_account_type,
                "date_range": (
                    f"{processing.system_date_range_start} to {processing.system_date_range_end}"
                    if processing.system_date_range_start
                    else None
                ),
            },
        }

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_sample_coa_excel():
    """Create a sample Excel file for Chart of Accounts upload."""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create headers
    headers = ["Code", "Name", "Type", "Tax Rate"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Sample data with Mongolian Cyrillic names
    sample_data = [
        ["1010", "Бэлэн мөнгө", "CURRENT_ASSET", ""],
        ["1020", "Банкны данс - Операцийн", "CURRENT_ASSET", ""],
        ["1030", "Дансны авлага", "CURRENT_ASSET", ""],
        ["1100", "Бараа материал", "INVENTORY", ""],
        ["1500", "Суурин хөрөнгө", "FIXED_ASSET", ""],
        ["2010", "Нийлүүлэгчийн өглөг", "CURRENT_LIABILITY", ""],
        ["2020", "Цалин хөлсний өглөг", "CURRENT_LIABILITY", ""],
        ["2100", "Банкны зээл", "NON_CURRENT_LIABILITY", ""],
        ["3010", "Хувь оролцоо", "EQUITY", ""],
        ["3020", "Олзын нөөц", "EQUITY", ""],
        ["4010", "Борлуулалтын орлого", "REVENUE", "Tax on Sales"],
        ["4020", "Бусад орлого", "OTHER_INCOME", ""],
        ["5010", "Худалдан авалтын зардал", "DIRECT_COST", "Tax on Purchases"],
        ["5020", "Цалин хөлс", "EXPENSE", ""],
        ["5030", "Захиргааны зардал", "OVERHEAD", ""],
        ["5040", "Элэгдэл", "DEPRECIATION", ""],
    ]
    
    # Add sample data
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    column_widths = [10, 30, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Save the file
    filename = "sample_chart_of_accounts.xlsx"
    wb.save(filename)
    
    print(f"✅ Sample Excel file created: {filename}")
    print(f"📍 File location: {os.path.abspath(filename)}")
    
    # Display file info
    print(f"\n📊 Sample file contains:")
    print(f"   • {len(sample_data)} sample accounts")
    print(f"   • Asset accounts: {sum(1 for row in sample_data if 'ASSET' in row[2])}")
    print(f"   • Liability accounts: {sum(1 for row in sample_data if 'LIABILITY' in row[2])}")
    print(f"   • Equity accounts: {sum(1 for row in sample_data if 'EQUITY' in row[2])}")
    print(f"   • Revenue accounts: {sum(1 for row in sample_data if 'REVENUE' in row[2] or 'INCOME' in row[2])}")
    print(f"   • Expense accounts: {sum(1 for row in sample_data if 'EXPENSE' in row[2] or 'COST' in row[2] or 'OVERHEAD' in row[2] or 'DEPRECIATION' in row[2])}")
    print(f"   • Accounts with tax rates: {sum(1 for row in sample_data if row[3])}")
    
    return filename

if __name__ == "__main__":
    create_sample_coa_excel()

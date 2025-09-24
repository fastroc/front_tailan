import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_restricted_codes_test_excel():
    """Create a test Excel file with some restricted codes to demonstrate validation."""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COA Test with Restricted"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create headers
    headers = ["Code", "Name", "Type", "Tax Rate"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Test data including RESTRICTED codes that should be blocked
    test_data = [
        # Valid accounts that should be allowed
        ["1010", "Бэлэн мөнгө - офис", "CURRENT_ASSET", ""],
        ["1030", "Дансны авлага", "CURRENT_ASSET", ""],
        ["4010", "Борлуулалтын орлого", "REVENUE", "Tax on Sales"],
        ["5020", "Цалин хөлс", "EXPENSE", ""],
        
        # RESTRICTED codes that should be BLOCKED
        ["111100", "Касст байгаа бэлэн мөнгө", "CURRENT_ASSET", ""],  # BLOCKED - Cash account
        ["112110", "Банканд байршуулсан харилцах", "CURRENT_ASSET", ""],  # BLOCKED - Bank account
        ["120000", "ЗЭЭЛ", "LIABILITY", ""],  # BLOCKED - Loan account
        ["150110", "ААНОАТ-ын авлага", "CURRENT_LIABILITY", ""],  # BLOCKED - Special receivable
        ["534190", "Банкны шимтгэлийн зардал", "EXPENSE", ""],  # BLOCKED - Bank charges
        
        # More valid accounts
        ["2010", "Нийлүүлэгчийн өглөг", "CURRENT_LIABILITY", ""],
        ["3010", "Хувь оролцоо", "EQUITY", ""],
        ["5030", "Захиргааны зардал", "OVERHEAD", ""],
    ]
    
    # Add test data
    for row_num, row_data in enumerate(test_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Highlight restricted rows in red
            if row_data[0] in ['111100', '112110', '120000', '150110', '534190']:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Adjust column widths
    column_widths = [10, 35, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Save the file
    filename = "test_restricted_accounts.xlsx"
    wb.save(filename)
    
    print(f"✅ Test Excel file created: {filename}")
    print(f"📍 File location: {os.path.abspath(filename)}")
    
    # Display file info
    print(f"\n📊 Test file contains:")
    print(f"   • {len(test_data)} total accounts")
    valid_accounts = [row for row in test_data if row[0] not in ['111100', '112110', '120000', '150110', '534190']]
    restricted_accounts = [row for row in test_data if row[0] in ['111100', '112110', '120000', '150110', '534190']]
    print(f"   • {len(valid_accounts)} valid accounts (should be accepted)")
    print(f"   • {len(restricted_accounts)} RESTRICTED accounts (should be BLOCKED):")
    
    for row in restricted_accounts:
        print(f"     - {row[0]}: {row[1]} (BLOCKED)")
    
    return filename

if __name__ == "__main__":
    create_restricted_codes_test_excel()
